# SAV Digital Environments
# Julian Kizanis
# generated in part by wxGlade 0.9.4 on Mon Nov 18 07:49:50 2019
#

from datetime import date
from getpass import getuser
from ntpath import basename

import wx
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import ColorScaleRule

# import xlsxwriter


# Declare GUI Constants
MENU_FILE_EXIT = wx.ID_ANY
DRAG_SOURCE = wx.ID_ANY

# Global Variables
pb = False

# Global Constants
ROUGH_PHASE = 1
FINISH_PHASE = 2
CONTINUE = 2
OVERRIDE = -2
CANCEL = -1


def check_for_duplicates(import_directory, imported_list):
    """This function checks if the file has already been imported"""
    for imp in imported_list:  # cycles through the directories of the previously imported files
        if import_directory == imp:  # if the directory we are trying to import matches a directory in the database
            return True  # we found a match
    return False  # no matches were found


def open_spreadsheet(directory):
    """This function tries to open a spreadsheet and prompts the user if it cannot"""
    while True:  # infinite loop
        try:
            dashboard = load_workbook(filename=directory, read_only=False, data_only=True)  # tries to open spreadsheet
            return dashboard  # returns the spreadsheet if it was opened
        except PermissionError:  # the spreadsheet is already open by something else
            dialog = DatasheetOpenDialog(basename(directory), None, wx.ID_ANY, "")  # asks if the user wants to retry
            retry = dialog.ShowModal()  # captures the users response
            if not retry:  # if the user doesn't want to retry
                return None


def append_dashboard(import_directories, phase, person, auto_replace, phase_check, dev_mode):
    """This function can import data using an external .xlsx map"""
    user = getuser()  # the current user
    if dev_mode == 1:
        if phase == ROUGH_PHASE:
            mappings_directory = "Dashboard Mappings Rough Testing.xlsx"
        else:
            mappings_directory = "Dashboard Mappings Testing.xlsx"
    else:
        if phase == ROUGH_PHASE:
            mappings_directory = f"C:/Users/{user}/SAV Digital Environments/SAV - Documents/Departments/Accounting/" \
                                 f"Job Costing/00 Master Job Costing Sheet/Job Costing Dashboard Import Program/" \
                                 f"Dashboard Mappings Rough.xlsx"
        else:
            mappings_directory = f"C:/Users/{user}/SAV Digital Environments/SAV - Documents/Departments/Accounting/" \
                                 f"Job Costing/00 Master Job Costing Sheet/Job Costing Dashboard Import Program/" \
                                 f"Dashboard Mappings.xlsx"

    map_book = open_spreadsheet(mappings_directory)  # contains the cell to cell mapping info
    if not map_book:  # checks if map_book is empty
        return None
    map_sheet = map_book.active  # finds the active spreadsheet
    if phase == ROUGH_PHASE:
        if dev_mode == 1:
            dashboard_directory = 'test dash rough.xlsx'
        else:
            dashboard_directory = f"C:/Users/{user}/SAV Digital Environments/SAV - Documents/Departments/Accounting/" \
                                  f"Job Costing/00 Master Job Costing Sheet/Job Costing_Master_Data_Sheet_Rough.xlsx"

        # creates a list containing the cell locations for the datasheet we are importing
        source_cells = []
        for cell in map_sheet['G4':'G38']:
            source_cells.append(cell[0].value)
        # creates a list containing the cell locations for the current master dashboard
        dashboard_columns = []
        for cell in map_sheet['F4':'F38']:
            dashboard_columns.append(cell[0].value)
        # creates a list that tells us what type of data we are importing
        phase_cells = []
        for cell in map_sheet['H4':'H38']:
            phase_cells.append(cell[0].value)
        number_formats = []
        for cell in map_sheet['I4':'I38']:
            number_formats.append(cell[0].value)
        functions = []
        for cell in map_sheet['J4':'J38']:
            functions.append(cell[0].value)
            print(cell[0].value)
        functions_filtered = []
        for cell in map_sheet['K4':'K38']:
            functions_filtered.append(cell[0].value)
            print(cell[0].value)
        left_borders = []
        for cell in map_sheet['L4':'L38']:
            left_borders.append(cell[0].value)
            print(cell[0].value)

    elif phase == FINISH_PHASE:
        if dev_mode == 1:
            dashboard_directory = 'test dash.xlsx'
        else:
            dashboard_directory = f"C:/Users/{user}/SAV Digital Environments/SAV - Documents/Departments/Accounting/" \
                                  f"Job Costing/00 Master Job Costing Sheet/Job Costing_Master_Data_Sheet.xlsx"

        # creates a list containing the cell locations for the datasheet we are importing
        source_cells = []
        for cell in map_sheet['G4':'G78']:
            source_cells.append(cell[0].value)
        # creates a list containing the cell locations for the current master dashboard
        dashboard_columns = []
        for cell in map_sheet['F4':'F78']:
            dashboard_columns.append(cell[0].value)
        # creates a list that tells us what type of data we are importing
        phase_cells = []
        for cell in map_sheet['H4':'H78']:
            phase_cells.append(cell[0].value)
        number_formats = []
        for cell in map_sheet['I4':'I78']:
            number_formats.append(cell[0].value)
        functions = []
        for cell in map_sheet['J4':'J78']:
            functions.append(cell[0].value)
            print(cell[0].value)
        functions_filtered = []
        for cell in map_sheet['K4':'K78']:
            functions_filtered.append(cell[0].value)
            print(cell[0].value)
        left_borders = []
        for cell in map_sheet['L4':'L78']:
            left_borders.append(cell[0].value)
            print(cell[0].value)

    # elif person == 'jake':
    #     dashboard_directory = f"{map_sheet['E2'].value}"
    #     source_cells = []
    #     for cell in map_sheet['G3':]:
    #         source_cells.append(cell[0].value)
    #     dashboard_column = []
    #     for cell in map_sheet['H3':]:
    #         dashboard_column.append(cell[0].value)

    else:
        return "code error: not valid export spreadsheet ID/Name"
    # dashboard_directory = dashboard_directory.format(user_name=user)
    print(person, ' directory: ', dashboard_directory)
    # opens the current master dashboard
    dashboard = open_spreadsheet(dashboard_directory)
    print("dashed")
    if not dashboard:
        return None

    if 'refresh' in person:
        temp_dirs = []
        if phase == ROUGH_PHASE:
            dir_column = 'AH'
        else:
            dir_column = 'BF'
        for index, temp_dir in enumerate(dashboard.active[dir_column]):
            if index > 4:
                if not temp_dir.value:
                    break
                print(basename(temp_dir.value))
                temp_dirs.append(temp_dir.value.replace('{user}', getuser()))
                if "Julian.Kizanis" in temp_dir.value:
                    temp_dir.value = temp_dir.value.replace('Julian.Kizanis', '{user}')
                if "cyndi.schoep" in temp_dir.value:
                    print('replacing...   ', temp_dir.value.replace('cyndi.schoep', '{user}'))
                    temp_dir.value = temp_dir.value.replace('cyndi.schoep', '{user}')
        import_directories = temp_dirs
        # for directory in import_directories:
        #     print(11111111111111111111, basename(directory))

    # addition row
    for dashboard_column, function, function_filtered in zip(dashboard_columns, functions, functions_filtered):
        # print(function)
        if function:
            dashboard.active[f'{dashboard_column}3'].value = function.strip('!')
        if function_filtered:
            dashboard.active[f'{dashboard_column}4'].value = function_filtered.strip('!')

    dashboard.save(dashboard_directory.format(user_name=user))

    for import_directory in import_directories:
        try:
            import_book = open_spreadsheet(import_directory)
        except FileNotFoundError:
            wx.MessageBox(f"{basename(import_directory)} is missing'. It cannot be imported!\n\n{import_directory}",
                          "Error", wx.OK | wx.ICON_INFORMATION)
            continue
        if not import_book:
            return None
        try:
            import_sheet = import_book['Data-Calculations']
        except KeyError:
            wx.MessageBox(f"{basename(import_directory)} does not have a sheet named 'Data-Calculations'. "
                          f"It cannot be imported!", "Error", wx.OK | wx.ICON_INFORMATION)
            return False

        change_row = 0
        last_row = 0
        for cell in dashboard.active['A']:  # searches through the project names in the master dashboard
            # print(person, cell.value, cell.row)
            date_column = ""
            for dashboard_column, phase_cell in zip(dashboard_columns, phase_cells):
                if phase_cell == 'rough_check':
                    date_column = dashboard_column
                if phase_cell == 'finish_check':
                    date_column = dashboard_column
            if cell.value == import_sheet['D2'].value:
                if (phase == ROUGH_PHASE or dashboard.active[f'{date_column}{cell.row}'].value) \
                        and auto_replace is False:
                    open_data_sheet = f"Name:{import_sheet['D2'].value}\nLocation:{import_sheet['D3'].value}\n" \
                                      f"PM:{import_sheet['D4'].value}\nDirectory:{import_directory}"
                    dialog = DatasheetAlreadyImportedDialog(open_data_sheet, dashboard.active[f'AU{cell.row}'].value,
                                                            dashboard.active[f'AV{cell.row}'].value,
                                                            None, wx.ID_ANY, "")

                    change_row = dialog.ShowModal()
                    print('change_row:  ', change_row)
                    if change_row == OVERRIDE:
                        change_row = cell.row
                        print('manual replace')
                    elif change_row == CANCEL:
                        return False
                elif (not dashboard.active[f'{date_column}{cell.row}'].value) or auto_replace is True:
                    change_row = cell.row
                    print('auto replace')
            if cell.value:
                last_row = cell.row

        if change_row == 0:
            if last_row == 0:
                return "data sheet error: first row empty"
            change_row = last_row + 1
            print(f"change_row:{change_row}")
        if change_row != CANCEL:
            print(person, phase, change_row)
            rough_complete = False
            finish_complete = False
            rough_ignore = finish_ignore = False
            # print('rough: ', source_cells, dashboard_columns, phase_cells, number_formats)
            for source_cell, dashboard_column, phase_cell, number_format, left_border in \
                    zip(source_cells, dashboard_columns, phase_cells, number_formats, left_borders):
                # print('cells', source_cell, dashboard_column, phase_cell)
                # print('dashboard_cell:  ', dashboard_column, ',', change_row)
                if phase_cell == 'Name':
                    dashboard.active[f'{dashboard_column}{change_row}'] = import_sheet[source_cell].value
                    dashboard.active[f'{dashboard_column}{change_row}'].alignment = Alignment(horizontal='left')

                if phase_cell == 'rough_check':
                    rough_complete = True

                    print("rough True: ", source_cell)
                    if not import_sheet[source_cell].value and phase_check is False:
                        print("No rough?  ", import_sheet[f'{dashboard_column}{change_row}'].value)
                        box = wx.MessageBox(f"Rough phase for {import_sheet['D2'].value} is not finished; "
                                            f"Do you want to import it anyways?", "Empty Import",
                                            wx.YES_NO | wx.ICON_INFORMATION)
                        if box != 2:
                            print(box)
                            rough_complete = False

                    if import_sheet[source_cell].value == 'N/A':
                        rough_ignore = True

                if 'rough' in phase_cell:
                    # print('rough: ', source_cell, dashboard_column, phase_cell, number_format)
                    if not rough_complete:
                        wx.MessageBox("Dashboard Mappings is incorrect, make sure rough_check is the first rough",
                                      "Empty Import", wx.OK | wx.ICON_INFORMATION)
                        return False
                    temp_cells = source_cell.split(' + ')
                    sum_cell = 0.0
                    for temp_index, temp_cell in enumerate(temp_cells):
                        temp_value = import_sheet[temp_cell].value
                        if temp_value is None:
                            temp_value = 0
                        if temp_index == 0:
                            sum_cell = temp_value
                        elif temp_value is not None:
                            sum_cell += temp_value
                    dashboard_cell = dashboard.active[f'{dashboard_column}{change_row}']
                    dashboard_cell.value = sum_cell

                    dashboard_cell.alignment = Alignment(horizontal='center')
                    if number_format:
                        dashboard_cell.number_format = number_format
                    else:
                        dashboard_cell.number_format = import_sheet[temp_cells[0]].number_format

                    if rough_ignore is True and (dashboard_cell.value == '#DIV/0!' or dashboard_cell.value == 0):
                        dashboard_cell.value = '#N/A'

                if phase_cell == 'finish_check' and phase == FINISH_PHASE:
                    finish_complete = True
                    if not import_sheet[source_cell].value and phase_check is False:
                        print("No finish?  ", import_sheet[f'{dashboard_column}{change_row}'].value)
                        box = wx.MessageBox(f"Finish phase for {import_sheet['D2'].value} is not finished; "
                                            f"Do you want to import it anyways?", "Empty Import",
                                            wx.YES_NO | wx.ICON_INFORMATION)
                        if box != 2:
                            print(box)
                            finish_complete = False
                    if import_sheet[source_cell].value == 'N/A':
                        finish_ignore = True

                if 'finish' in phase_cell and phase == FINISH_PHASE and finish_complete:
                    # print('finish: ', source_cell, dashboard_column, phase_cell, number_format)
                    temp_cells = source_cell.split(' + ')
                    sum_cell = 0.0
                    # print('temp_cells =', temp_cells)
                    for temp_index, temp_cell in enumerate(temp_cells):
                        temp_value = import_sheet[temp_cell].value
                        if temp_value is None:
                            temp_value = 0
                        # print('temp_cell =', temp_cell,
                        #       '\ttemp_value =', temp_value,
                        #       '\tsum_cell =', sum_cell)
                        if temp_index == 0:
                            sum_cell = temp_value
                        else:
                            # print(f"{dashboard_column} sum_cell = {sum_cell} += {temp_value}")
                            sum_cell += temp_value
                    dashboard_cell = dashboard.active[f'{dashboard_column}{change_row}']
                    dashboard_cell.value = sum_cell

                    dashboard_cell.alignment = Alignment(horizontal='center')
                    if number_format:
                        dashboard_cell.number_format = number_format
                    else:
                        dashboard_cell.number_format = import_sheet[temp_cells[0]].number_format

                    if finish_ignore is True and (dashboard_cell.value == '#DIV/0!' or dashboard_cell.value == 0):
                        dashboard_cell.value = '#N/A'

                if phase_cell == 'formula':
                    dashboard_cell = dashboard.active[f'{dashboard_column}{change_row}']
                    dashboard_cell.value = source_cell.strip('!').replace("{change_row}", str(change_row))
                    dashboard_cell.alignment = Alignment(horizontal='center')
                    dashboard_cell.number_format = number_format

                if phase_cell == 'logging':
                    dashboard_cell = dashboard.active[f'{dashboard_column}{change_row}']
                    print('dash_cell: ', dashboard_cell, 'source_cell', source_cell, 'imp_dir: ', import_directory)
                    if source_cell == 'directory':
                        print('logging directory')
                        dashboard_cell.value = import_directory.replace(getuser(), '{user}')

                    elif source_cell == 'user':
                        print('logging user')
                        dashboard_cell.value = user
                    elif source_cell == 'date':
                        print('logging date')
                        dashboard_cell.value = date.today()

                if phase_cell == 'format':
                    split_format_cells = source_cell.split(', ')
                    start_color = str(split_format_cells[0].split()[0])
                    start_value = split_format_cells[0].split()[1]
                    mid_color = str(split_format_cells[1].split()[0])
                    mid_value = split_format_cells[1].split()[1]
                    end_color = str(split_format_cells[2].split()[0])
                    end_value = split_format_cells[2].split()[1]

                    split_location_cells = dashboard_column.split(', ')
                    input_column = split_location_cells[0]
                    output_column = split_location_cells[1]

                    if input_column == output_column:
                        print(start_value, start_color, mid_value, mid_color, end_value, end_color)
                        format_rule = ColorScaleRule(start_type='num', start_value=start_value, start_color=start_color,
                                                     mid_type='num', mid_value=mid_value, mid_color=mid_color,
                                                     end_type='num', end_value=end_value, end_color=end_color)
                        dashboard.active.conditional_formatting.add(f'{output_column}{change_row}', format_rule)
                    else:
                        input_value = dashboard.active[f'{input_column}{change_row}'].value
                        output_cell = dashboard.active[f'{output_column}{change_row}']
                        print(input_value, start_value, end_value)
                        print('input_value', input_value, 'start_value', start_value)
                        try:
                            if float(input_value) < float(start_value):
                                output_cell.fill = PatternFill(fgColor=start_color, fill_type='solid')
                            elif float(input_value) > float(end_value):
                                output_cell.fill = PatternFill(fgColor=end_color, fill_type='solid')
                            else:
                                output_cell.fill = PatternFill(fgColor=mid_color, fill_type='solid')
                        except ValueError as e:
                            wx.MessageBox(str(e), "ValueError", wx.OK | wx.ICON_INFORMATION)
                        except TypeError:
                            print('empty')

                if left_border == 'thick':
                    dashboard_cell = dashboard.active[f'{dashboard_column}{change_row}']
                    dashboard_cell.border = Border(left=Side(style='medium'),
                                                   right=Side(style='thin'),
                                                   top=Side(style='thin'),
                                                   bottom=Side(style='thin'))

                if left_border == 'thin':
                    dashboard_cell = dashboard.active[f'{dashboard_column}{change_row}']
                    dashboard_cell.border = Border(left=Side(style='thin'),
                                                   right=Side(style='thin'),
                                                   top=Side(style='thin'),
                                                   bottom=Side(style='thin'))

    print('saving to: ', dashboard_directory.format(user_name=user))
    dashboard.save(dashboard_directory.format(user_name=user))

    # dashboard_link_book = xlsxwriter.Workbook('test dash.xlsx')
    # for dashboard_link_sheet in dashboard_link_book.worksheets():
    #     forward_slashed_directory = import_directory_test.replace('\\', '/')
    #     dashboard_link_sheet.write_url(f'A{change_row}',
    #                                    f"external:{forward_slashed_directory}",
    #                                    string=str(dashboard.active[f'A{change_row}'].value))

    return True


class FileDropTarget(wx.FileDropTarget):
    """ This object implements Drop Target functionality for Files """

    def __init__(self, obj, import_files):
        """ Initialize the Drop Target, passing in the Object Reference to
        indicate what should receive the dropped files """
        # Initialize the wxFileDropTarget Object
        wx.FileDropTarget.__init__(self)
        # Store the Object Reference for dropped files
        self.obj = obj
        self.import_files = import_files

    def OnDropFiles(self, x, y, file_names):
        """ Implement File Drop """
        # Move Insertion Point to the end of the widget's text
        self.obj.SetInsertionPointEnd()
        # append a list of the file names dropped
        dup_check = False
        for file in file_names:
            for iFile in self.import_files:
                if file == iFile:
                    dup_check = True
            if not (file.endswith('.xlsx') or file.endswith('.xlsm')):
                wx.MessageBox("Incorrect file type. Please choose an .xlsx or an .xlsm file.",
                              "Error", wx.OK | wx.ICON_INFORMATION)
                continue
            if not dup_check:
                self.obj.WriteText(basename(file) + '\n')
                self.import_files.append(file)
            else:
                print("Removed duplicate import file!")
                wx.MessageBox("File already in import list.", "Error", wx.OK | wx.ICON_INFORMATION)
                dup_check = False
        self.obj.WriteText('\n')
        return True


class FirstFrame(wx.Frame):
    """This object is the main window"""

    def __init__(self, *args, **kwds):
        kwds["style"] = kwds.get("style", 0) | wx.DEFAULT_FRAME_STYLE
        wx.Frame.__init__(self, *args, **kwds)

        self.import_files = []

        self.SetSize((640, 428))
        self.button_browse = wx.FilePickerCtrl(self)
        # self.button_4.Bind(wx.EVT_FILEPICKER_CHANGED, self.on_choose_file)
        self.text_ctrl_drag_drop = wx.TextCtrl(self, wx.ID_ANY, "", style=wx.TE_MULTILINE | wx.TE_READONLY)

        drop_target = FileDropTarget(self.text_ctrl_drag_drop, self.import_files)

        # Link the Drop Target Object to the Text Control
        self.text_ctrl_drag_drop.SetDropTarget(drop_target)

        # initializes the buttons
        self.choice_phase = wx.Choice(self, wx.ID_ANY, choices=["Choose Phase", "Rough In", "Finish"])
        self.checkbox_auto_replace = wx.CheckBox(self, wx.ID_ANY, "Automatically Replace")
        self.checkbox_phase_check = wx.CheckBox(self, wx.ID_ANY, "Check for Phase Completion Date")
        self.checkbox_testing = wx.CheckBox(self, wx.ID_ANY, "Testing/Development Mode")
        self.panel_1 = wx.Panel(self, wx.ID_ANY)
        self.button_continue = wx.Button(self, wx.ID_ANY, "Continue")
        self.button_cancel = wx.Button(self, wx.ID_ANY, "Cancel")
        self.button_clear = wx.Button(self, wx.ID_ANY, "Clear")
        self.button_refresh = wx.Button(self, wx.ID_ANY, "Refresh")

        self.__set_properties()
        self.__do_layout()
        self.SetMinSize((440, 370))

        # initializes the events
        self.Bind(wx.EVT_FILEPICKER_CHANGED, self.on_choose_file, self.button_browse)
        self.Bind(wx.EVT_CHOICE, self.on_phase_selection, self.choice_phase)
        self.Bind(wx.EVT_CHECKBOX, self.on_auto_replace_checkbox, self.checkbox_auto_replace)
        self.Bind(wx.EVT_CHECKBOX, self.on_phase_check_checkbox, self.checkbox_phase_check)
        self.Bind(wx.EVT_CHECKBOX, self.on_testing_checkbox, self.checkbox_testing)
        self.Bind(wx.EVT_BUTTON, self.on_continue_from_main_window, self.button_continue)
        self.Bind(wx.EVT_BUTTON, self.on_cancel_program, self.button_cancel)
        self.Bind(wx.EVT_BUTTON, self.on_clear, self.button_clear)
        self.Bind(wx.EVT_BUTTON, self.on_refresh, self.button_refresh)
        self.Bind(wx.EVT_ICONIZE, self.on_minimize)

    def __set_properties(self):
        self.SetTitle("Import Project Datasheet")
        _icon = wx.NullIcon
        _icon.CopyFromBitmap(wx.Bitmap("SAV-Social-Profile.jpg", wx.BITMAP_TYPE_ANY))
        self.SetIcon(_icon)

        self.SetBackgroundColour(wx.Colour(255, 255, 255))
        self.choice_phase.SetMinSize((102, 23))
        self.choice_phase.SetSelection(2)
        self.checkbox_auto_replace.SetValue(1)
        self.checkbox_phase_check.SetValue(1)
        self.checkbox_testing.SetValue(0)

    def __do_layout(self):
        sizer_5 = wx.BoxSizer(wx.VERTICAL)
        sizer_9 = wx.GridBagSizer(0, 0)
        sizer_6 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_8 = wx.BoxSizer(wx.VERTICAL)
        sizer_11 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_15 = wx.BoxSizer(wx.VERTICAL)
        sizer_12 = wx.BoxSizer(wx.VERTICAL)
        sizer_13 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_7 = wx.BoxSizer(wx.VERTICAL)
        sizer_14 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_16 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_10 = wx.BoxSizer(wx.VERTICAL)

        label_1 = wx.StaticText(self, wx.ID_ANY, f"Hello {getuser()}! This program imports job costing spreadsheets "
                                                 f"to a master dashboard.")
        sizer_10.Add(label_1, 0, wx.ALL, 5)
        static_line_1 = wx.StaticLine(self, wx.ID_ANY)
        sizer_10.Add(static_line_1, 0, wx.EXPAND, 0)
        sizer_5.Add(sizer_10, 0, wx.EXPAND, 0)
        sizer_16.Add(self.button_browse, 0, wx.ALL, 12)
        label_6 = wx.StaticText(self, wx.ID_ANY, "Or drag and drop files below")
        sizer_16.Add(label_6, 0, wx.ALIGN_CENTER, 0)
        sizer_7.Add(sizer_16, 0, wx.EXPAND, 0)
        sizer_14.Add(self.text_ctrl_drag_drop, 1, wx.EXPAND, 0)
        sizer_7.Add(sizer_14, 1, wx.EXPAND, 0)
        sizer_6.Add(sizer_7, 2, wx.EXPAND, 0)
        bitmap_2 = wx.StaticBitmap(self, wx.ID_ANY, wx.Bitmap("SAV-Company-Logo.png", wx.BITMAP_TYPE_ANY))
        sizer_12.Add(bitmap_2, 0, wx.BOTTOM | wx.RIGHT | wx.TOP, 12)
        sizer_13.Add(self.choice_phase, 0, wx.BOTTOM | wx.LEFT, 6)
        sizer_12.Add(sizer_13, 1, wx.EXPAND, 0)
        sizer_8.Add(sizer_12, 0, wx.EXPAND, 0)
        sizer_15.Add(self.checkbox_auto_replace, 0, wx.LEFT | wx.RIGHT | wx.TOP, 6)
        sizer_15.Add(self.checkbox_phase_check, 0, wx.LEFT | wx.RIGHT | wx.TOP, 6)
        sizer_15.Add(self.checkbox_testing, 0, wx.LEFT | wx.RIGHT | wx.TOP, 6)
        sizer_11.Add(sizer_15, 1, wx.EXPAND, 0)
        sizer_8.Add(sizer_11, 1, wx.EXPAND, 0)
        sizer_6.Add(sizer_8, 0, wx.EXPAND | wx.LEFT, 6)
        sizer_5.Add(sizer_6, 1, wx.EXPAND, 0)
        sizer_9.Add(self.panel_1, (0, 0), (1, 1), wx.EXPAND, 0)
        sizer_9.Add(self.button_continue, (0, 1), (1, 1), wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL, 6)
        sizer_9.Add(self.button_cancel, (0, 4), (1, 1), wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL, 6)
        sizer_9.Add(self.button_clear, (0, 3), (1, 1), wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL, 6)
        sizer_9.Add(self.button_refresh, (0, 2), (1, 1), wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL, 6)
        sizer_5.Add(sizer_9, 0, wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL | wx.EXPAND, 12)
        self.SetSizer(sizer_5)
        self.Layout()

    def on_choose_file(self, event):  # button_browse
        dup_check = False
        file = self.button_browse.GetPath()  # catches what file the user chose
        for iFile in self.import_files:  # checks if file is already in the to-be imported list
            if file == iFile:
                dup_check = True
        if not file.endswith('.xlsx'):
            wx.MessageBox("Incorrect file type. Please choose an .xlsx file.", "Error", wx.OK | wx.ICON_INFORMATION)
            event.skip()
        if not dup_check:
            self.import_files.append(file)
            self.text_ctrl_drag_drop.WriteText(basename(file) + '\n')  # shows the user they chose this
        else:
            print("Removed duplicate import file!")
            wx.MessageBox("File already in import list.", "Error", wx.OK | wx.ICON_INFORMATION)
        event.Skip()

    def on_phase_selection(self, event):  # event handler
        print(self.choice_phase.GetSelection())
        event.Skip()

    def on_auto_replace_checkbox(self, event):  # event handler
        print(self.checkbox_auto_replace.GetValue())
        event.Skip()

    def on_phase_check_checkbox(self, event):  # event handler
        print(self.checkbox_phase_check.GetValue())
        event.Skip()

    def on_testing_checkbox(self, event):  # event handler
        print(self.checkbox_testing.GetValue())
        event.Skip()

    def on_continue_from_main_window(self, event):  # event handler
        if self.choice_phase.GetSelection() == 0:  # no phase was chosen
            wx.MessageBox("Please choose a phase.", "Error", wx.OK | wx.ICON_INFORMATION)
        elif not self.import_files:
            wx.MessageBox("Please choose a file to import.", "Error", wx.OK | wx.ICON_INFORMATION)
        else:
            # for tracking if something went wrong
            default_check = kacey_check = True
            if self.checkbox_testing.GetValue() is False:
                try:
                    # default_check = append_dashboard(self.import_files, self.choice_phase.GetSelection(), 'default',
                    #                                  self.checkbox_auto_replace.GetValue(),
                    #                                  self.checkbox_phase_check.GetValue())

                    kacey_check = append_dashboard(self.import_files, self.choice_phase.GetSelection(), 'kacey',
                                                   self.checkbox_auto_replace.GetValue(),
                                                   self.checkbox_phase_check.GetValue(),
                                                   self.checkbox_testing.GetValue())
                except Exception as e:
                    wx.MessageBox(str(e), "Error!", wx.OK | wx.ICON_INFORMATION)
            else:
                kacey_check = append_dashboard(self.import_files, self.choice_phase.GetSelection(), 'kacey',
                                               self.checkbox_auto_replace.GetValue(),
                                               self.checkbox_phase_check.GetValue(),
                                               self.checkbox_testing.GetValue())

            if default_check is True and kacey_check is True:  # if everything was successfully imported
                wx.MessageBox(f"{self.text_ctrl_drag_drop.GetValue()}\n Was successfully imported!", "Done!",
                              wx.OK | wx.ICON_INFORMATION)
            else:
                wx.MessageBox("Something went wrong or did not import", "Done!", wx.OK | wx.ICON_INFORMATION)
            self.text_ctrl_drag_drop.SetValue("")  # resets the program
            self.import_files.clear()  # resets the program

        event.Skip()

    def on_cancel_program(self, event):  # event handler
        print(getuser())
        self.Destroy()
        event.Skip()

    def on_clear(self, event):  # resets the program
        self.text_ctrl_drag_drop.SetValue("")
        global pb
        self.import_files.clear()
        pb = not pb
        event.Skip()

    def on_refresh(self, event):  # Refreshes the data
        # wx.MessageBox("Feature not finished yet", "Refreshing!", wx.ICON_INFORMATION)
        # return False
        # TODO
        # refresh_box = wx.MessageBox("Refreshing the master dashboards...", "Refreshing!", wx.ICON_INFORMATION)
        default_check = kacey_check = False
        if self.checkbox_testing.GetValue() is False:
            # try:
            # default_check = refresh_dashboard(refresh_box, 'default', self.checkbox_phase_check.GetValue())
            kacey_check = append_dashboard(self.import_files, self.choice_phase.GetSelection(), 'kacey refresh',
                                           self.checkbox_auto_replace.GetValue(),
                                           self.checkbox_phase_check.GetValue(),
                                           self.checkbox_testing.GetValue())
            default_check = True
            # except Exception as e:
            #     wx.MessageBox(str(e), "Error!", wx.OK | wx.ICON_INFORMATION)
        else:
            kacey_check = append_dashboard(self.import_files, self.choice_phase.GetSelection(), 'kacey refresh',
                                           self.checkbox_auto_replace.GetValue(),
                                           self.checkbox_phase_check.GetValue(),
                                           self.checkbox_testing.GetValue())
            default_check = True

        if default_check is True and kacey_check is True:
            wx.MessageBox("Done!", "Refreshing!", wx.ICON_INFORMATION)
        else:
            wx.MessageBox("Error!", "Refreshing!", wx.ICON_INFORMATION)

        event.Skip()

    @staticmethod
    def on_minimize(event):  # easter egg
        global pb
        if pb:
            wx.MessageBox("Or is it Peanut butter?", "Peanutbutter!", wx.OK | wx.ICON_INFORMATION)
        pb = False
        event.Skip()


class DatasheetOpenDialog(wx.Dialog):
    def __init__(self, open_data_sheet, *args, **kwds):
        kwds["style"] = kwds.get("style", 0) | wx.DEFAULT_DIALOG_STYLE
        wx.Dialog.__init__(self, *args, **kwds)
        self.open_data_sheet = open_data_sheet
        self.text_ctrl_open_datasheet = wx.TextCtrl(self, wx.ID_ANY,
                                                    f"{open_data_sheet} is open by a user. Please close "
                                                    f"{open_data_sheet} and click \"Retry\".",
                                                    style=wx.BORDER_NONE | wx.TE_MULTILINE | wx.TE_READONLY)
        self.panel_2 = wx.Panel(self, wx.ID_ANY)
        self.button_1 = wx.Button(self, wx.ID_ANY, "Retry")
        self.button_5 = wx.Button(self, wx.ID_ANY, "Back")

        self.__set_properties()
        self.__do_layout()

        self.Bind(wx.EVT_TEXT, self.text_ctrl_open_data_sheet, self.text_ctrl_open_datasheet)
        self.Bind(wx.EVT_BUTTON, self.on_retry, self.button_1)
        self.Bind(wx.EVT_BUTTON, self.on_back, self.button_5)

    def __set_properties(self):
        _icon = wx.NullIcon
        _icon.CopyFromBitmap(wx.Bitmap("SAV-Social-Profile.jpg", wx.BITMAP_TYPE_ANY))
        self.SetIcon(_icon)

        self.SetTitle("dialog_3")
        self.text_ctrl_open_datasheet.SetBackgroundColour(wx.Colour(255, 255, 255))

    def __do_layout(self):
        sizer_1 = wx.BoxSizer(wx.VERTICAL)
        sizer_2 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_1.Add(self.text_ctrl_open_datasheet, 0, wx.ALL | wx.EXPAND, 12)
        sizer_2.Add(self.panel_2, 1, 0, 0)
        sizer_2.Add(self.button_1, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_2.Add(self.button_5, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_1.Add(sizer_2, 1, wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL | wx.EXPAND | wx.FIXED_MINSIZE, 1)
        self.SetSizer(sizer_1)
        sizer_1.Fit(self)
        self.Layout()

    def text_ctrl_open_data_sheet(self, event):  # event handler
        print(f"{self.open_data_sheet} is currently open by a user!")
        event.Skip()

    def on_retry(self, event):  # event handler
        self.EndModal(True)
        self.Destroy()
        event.Skip()

    def on_back(self, event):  # event handler
        self.EndModal(False)
        self.Destroy()
        event.Skip()


class DatasheetAlreadyImportedDialog(wx.Dialog):
    def __init__(self, open_sheet, imported_by, imported_date, *args, **kwds):
        kwds["style"] = kwds.get("style", 0) | wx.DEFAULT_DIALOG_STYLE
        wx.Dialog.__init__(self, *args, **kwds)
        self.open_data_sheet = open_sheet
        self.text_ctrl_already_imported = wx.TextCtrl(self, wx.ID_ANY,
                                                      f"{open_sheet}\nHas already been imported by {imported_by} on "
                                                      f"{imported_date}. If you would  like to import the project as a "
                                                      f"new project, select \"Duplicate\". If you want to refresh the "
                                                      f"existing data, select \"Replace\".",
                                                      style=wx.BORDER_NONE | wx.TE_MULTILINE | wx.TE_READONLY)
        self.panel_2 = wx.Panel(self, wx.ID_ANY)
        self.button_6 = wx.Button(self, wx.ID_ANY, "Duplicate")
        self.button_1 = wx.Button(self, wx.ID_ANY, "Replace")
        self.button_5 = wx.Button(self, wx.ID_ANY, "Back")

        self.__set_properties()
        self.__do_layout()

        self.Bind(wx.EVT_TEXT, self.text_ctrl_open_data_sheet, self.text_ctrl_already_imported)
        self.Bind(wx.EVT_BUTTON, self.on_duplicate, self.button_6)
        self.Bind(wx.EVT_BUTTON, self.on_replace, self.button_1)
        self.Bind(wx.EVT_BUTTON, self.on_back, self.button_5)

    def __set_properties(self):
        _icon = wx.NullIcon
        _icon.CopyFromBitmap(wx.Bitmap("SAV-Social-Profile.jpg", wx.BITMAP_TYPE_ANY))
        self.SetIcon(_icon)

        self.SetTitle("dialog_2")
        self.text_ctrl_already_imported.SetBackgroundColour(wx.Colour(255, 255, 255))

    def __do_layout(self):
        sizer_1 = wx.BoxSizer(wx.VERTICAL)
        sizer_2 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_1.Add(self.text_ctrl_already_imported, 0, wx.ALL | wx.EXPAND, 12)
        sizer_2.Add(self.panel_2, 1, 0, 0)
        sizer_2.Add(self.button_6, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_2.Add(self.button_1, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_2.Add(self.button_5, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_1.Add(sizer_2, 1, wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL | wx.EXPAND | wx.FIXED_MINSIZE, 1)
        self.SetSizer(sizer_1)
        sizer_1.Fit(self)
        self.Layout()

    def text_ctrl_open_data_sheet(self, event):  # event handler
        print(f"{self.open_data_sheet} has already been imported")
        event.Skip()

    def on_duplicate(self, event):  # event handler
        if getuser() == "Julian.Kizanis":
            dialog = AreYouSureDuplicateDialog(None, wx.ID_ANY, "")
            answer = dialog.ShowModal()
            if answer:
                self.EndModal(0)
            else:
                self.EndModal(CANCEL)
            self.Destroy()
        else:
            wx.MessageBox("This functionality has been disabled, please contact "
                          "Julian.Kizanis if you wish to duplicate project entries.",
                          "Duplicate", wx.OK | wx.ICON_INFORMATION)
        event.Skip()

    def on_replace(self, event):  # event handler
        dialog = AreYouSureReplaceDialog(None, wx.ID_ANY, "")
        answer = dialog.ShowModal()
        if answer:
            self.EndModal(OVERRIDE)
        else:
            self.EndModal(CANCEL)
        self.Destroy()
        event.Skip()

    def on_back(self, event):  # event handler
        self.EndModal(CANCEL)
        self.Destroy()
        event.Skip()


class AreYouSureReplaceDialog(wx.Dialog):
    def __init__(self, *args, **kwds):
        kwds["style"] = kwds.get("style", 0) | wx.DEFAULT_DIALOG_STYLE
        wx.Dialog.__init__(self, *args, **kwds)
        self.panel_2 = wx.Panel(self, wx.ID_ANY)
        self.button_1 = wx.Button(self, wx.ID_ANY, "Replace")
        self.button_5 = wx.Button(self, wx.ID_ANY, "Back")

        self.__set_properties()
        self.__do_layout()

        self.Bind(wx.EVT_BUTTON, self.on_replace, self.button_1)
        self.Bind(wx.EVT_BUTTON, self.on_back, self.button_5)

    def __set_properties(self):
        self.SetTitle("dialog")
        _icon = wx.NullIcon
        _icon.CopyFromBitmap(wx.Bitmap("SAV-Social-Profile.jpg", wx.BITMAP_TYPE_ANY))
        self.SetIcon(_icon)

    def __do_layout(self):
        sizer_1 = wx.BoxSizer(wx.VERTICAL)
        sizer_2 = wx.BoxSizer(wx.HORIZONTAL)
        label_2 = wx.StaticText(self, wx.ID_ANY,
                                "Are you Sure you want to replace/overwrite the project? "
                                "The old data will not be saved.")
        label_2.Wrap(300)
        sizer_1.Add(label_2, 0, wx.ALL, 12)
        sizer_2.Add(self.panel_2, 1, 0, 0)
        sizer_2.Add(self.button_1, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_2.Add(self.button_5, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_1.Add(sizer_2, 1, wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL | wx.EXPAND | wx.FIXED_MINSIZE, 1)
        self.SetSizer(sizer_1)
        sizer_1.Fit(self)
        self.Layout()

    def on_replace(self, event):  # event handler
        self.EndModal(True)
        self.Destroy()
        event.Skip()

    def on_back(self, event):  # event handler
        self.EndModal(False)
        self.Destroy()
        event.Skip()


class AreYouSureDuplicateDialog(wx.Dialog):
    def __init__(self, *args, **kwds):
        kwds["style"] = kwds.get("style", 0) | wx.DEFAULT_DIALOG_STYLE
        wx.Dialog.__init__(self, *args, **kwds)
        self.panel_2 = wx.Panel(self, wx.ID_ANY)
        self.button_1 = wx.Button(self, wx.ID_ANY, "Duplicate")
        self.button_5 = wx.Button(self, wx.ID_ANY, "Back")

        self.__set_properties()
        self.__do_layout()

        self.Bind(wx.EVT_BUTTON, self.on_duplicate, self.button_1)
        self.Bind(wx.EVT_BUTTON, self.on_back, self.button_5)

    def __set_properties(self):
        _icon = wx.NullIcon
        _icon.CopyFromBitmap(wx.Bitmap("SAV-Social-Profile.jpg", wx.BITMAP_TYPE_ANY))
        self.SetIcon(_icon)
        self.SetTitle("dialog_1")

    def __do_layout(self):
        sizer_1 = wx.BoxSizer(wx.VERTICAL)
        sizer_2 = wx.BoxSizer(wx.HORIZONTAL)
        label_2 = wx.StaticText(self, wx.ID_ANY, "Are you Sure you want to add the project as a duplicate?")
        label_2.Wrap(300)
        sizer_1.Add(label_2, 0, wx.ALL, 12)
        sizer_2.Add(self.panel_2, 1, 0, 0)
        sizer_2.Add(self.button_1, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_2.Add(self.button_5, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_1.Add(sizer_2, 1, wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL | wx.EXPAND | wx.FIXED_MINSIZE, 1)
        self.SetSizer(sizer_1)
        sizer_1.Fit(self)
        self.Layout()

    def on_duplicate(self, event):  # event handler
        self.EndModal(True)
        self.Destroy()
        event.Skip()

    def on_back(self, event):  # event handler
        self.EndModal(False)
        self.Destroy()
        event.Skip()


class SuccessFrame(wx.Frame):
    def __init__(self, *args, **kwds):
        kwds["style"] = kwds.get("style", 0) | wx.DEFAULT_FRAME_STYLE | wx.STAY_ON_TOP
        wx.Frame.__init__(self, *args, **kwds)
        self.SetSize((350, 150))
        self.panel_2 = wx.Panel(self, wx.ID_ANY)
        self.button_5 = wx.Button(self, wx.ID_ANY, "Okay")

        self.__set_properties()
        self.__do_layout()

        self.Bind(wx.EVT_BUTTON, self.on_okay, self.button_5)

    def __set_properties(self):
        _icon = wx.NullIcon
        _icon.CopyFromBitmap(wx.Bitmap("SAV-Social-Profile.jpg", wx.BITMAP_TYPE_ANY))
        self.SetIcon(_icon)

        self.SetTitle("frame_2")
        self.SetBackgroundColour(wx.Colour(255, 255, 255))

    def __do_layout(self):
        sizer_1 = wx.BoxSizer(wx.VERTICAL)
        sizer_2 = wx.BoxSizer(wx.HORIZONTAL)
        label_2 = wx.StaticText(self, wx.ID_ANY, "The project was successfully imported!")
        label_2.Wrap(300)
        sizer_1.Add(label_2, 0, wx.ALL, 12)
        sizer_2.Add(self.panel_2, 1, 0, 0)
        sizer_2.Add(self.button_5, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_1.Add(sizer_2, 1, wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL | wx.EXPAND | wx.FIXED_MINSIZE, 1)
        self.SetSizer(sizer_1)
        self.Layout()

    def on_okay(self, event):  # event handler
        print("Event handler 'on_okay' not implemented!")
        self.Destroy()
        event.Skip()


class ErrorFrame(wx.Frame):
    def __init__(self, *args, **kwds):
        kwds["style"] = kwds.get("style", 0) | wx.DEFAULT_FRAME_STYLE | wx.STAY_ON_TOP
        wx.Frame.__init__(self, *args, **kwds)
        self.SetSize((350, 150))
        self.panel_2 = wx.Panel(self, wx.ID_ANY)
        self.button_5 = wx.Button(self, wx.ID_ANY, "Okay")

        self.__set_properties()
        self.__do_layout()

        self.Bind(wx.EVT_BUTTON, self.on_okay, self.button_5)

    def __set_properties(self):
        _icon = wx.NullIcon
        _icon.CopyFromBitmap(wx.Bitmap("SAV-Social-Profile.jpg", wx.BITMAP_TYPE_ANY))
        self.SetIcon(_icon)
        self.SetTitle("frame_2")
        self.SetBackgroundColour(wx.Colour(255, 255, 255))

    def __do_layout(self):
        sizer_1 = wx.BoxSizer(wx.VERTICAL)
        sizer_2 = wx.BoxSizer(wx.HORIZONTAL)
        label_2 = wx.StaticText(self, wx.ID_ANY, "An unexpected error has occurred!")
        label_2.Wrap(300)
        sizer_1.Add(label_2, 0, wx.ALL, 12)
        sizer_2.Add(self.panel_2, 1, 0, 0)
        sizer_2.Add(self.button_5, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_1.Add(sizer_2, 1, wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL | wx.EXPAND | wx.FIXED_MINSIZE, 1)
        self.SetSizer(sizer_1)
        self.Layout()

    @staticmethod
    def on_okay(event):  # event handler
        print("Event handler 'on_okay' not implemented!")
        event.Skip()


class MyApp(wx.App):
    def OnInit(self):
        self.frame = FirstFrame(None, wx.ID_ANY, "")
        self.SetTopWindow(self.frame)
        self.frame.Show()
        return True


if __name__ == "__main__":
    ImportProjectDatasheets = MyApp(0)
    ImportProjectDatasheets.MainLoop()
