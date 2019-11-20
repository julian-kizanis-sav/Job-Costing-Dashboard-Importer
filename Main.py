# SAV Digital Environments
# Julian Kizanis
print("Commpany:\tSAV Digital Environments\nDeveloper:\tJulian Kizanis\n\
Powered By:\tAnaconda\n\n")
#
# generated by wxGlade 0.9.4 on Mon Nov 18 07:49:50 2019
#

import wx
from openpyxl import load_workbook

# Declare GUI Constants
MENU_FILE_EXIT = wx.ID_ANY
DRAG_SOURCE = wx.ID_ANY

# Global Variables
ImportFiles = []
openDataSheet = "None"
changeRow = 0

# Global Constants
ROUGH_PHASE = 1
FINISH_PHASE = 2


def check_for_dups(importDirectory, ImportedList):
    for imp in ImportedList:
        if importDirectory == imp:
            return True
    return False


def append_kaceys_dashboard(impDir, phase):
    dashDir = r"C:\Users\Julian.Kizanis\SAV Digital Environments\SAV - Documents\Departments\Accounting\Job Costing\00 Master Job Costing Sheet\Job Costing_Master_Data_Sheet.xlsx"
    # dashDir = "testing.xlsx"
    dashboard = load_workbook(filename=dashDir, read_only=False, data_only=True)
    for imp in impDir:
        impBook = load_workbook(filename=imp, read_only=False, data_only=True)
        impSheet = impBook['Data-Calculations']

        global changeRow
        changeRow = 0
        for cell in dashboard.active['AT']:
            print(cell.value)
            if cell.value == imp and (phase == ROUGH_PHASE or dashboard.active[f'E{cell.row}']):
                global openDataSheet
                openDataSheet = f"Name:{impSheet['D2'].value}\nLocation:{impSheet['D3'].value}\nPM:{impSheet['D4'].value}\nDirectory:{imp}"
                changeRow = FirstFrame.duplicate_import(cell.row, imp)
            if cell.value:
                lastRow = cell.row

        if changeRow == 0:
            changeRow = lastRow + 1
            print(f"ChangeRow:{changeRow}")
        if changeRow > -1:
            if not impSheet['D5'].value:
                print("need to implement pop up for importing an empty finish phase")
            else:
                dashboard.active[f'A{changeRow}'].value = impSheet['D3'].value
                dashboard.active[f'B{changeRow}'].value = impSheet['D2'].value
                dashboard.active[f'C{changeRow}'].value = impSheet['D4'].value
                dashboard.active[f'D{changeRow}'].value = impSheet['D5'].value
                dashboard.active[f'R{changeRow}'].value = impSheet['N13'].value
                dashboard.active[f'S{changeRow}'].value = impSheet['O13'].value
                dashboard.active[f'T{changeRow}'].value = impSheet['P13'].value
                dashboard.active[f'X{changeRow}'].value = impSheet['D47'].value
                dashboard.active[f'Y{changeRow}'].value = impSheet['E47'].value
                dashboard.active[f'Z{changeRow}'].value = impSheet['F47'].value
                dashboard.active[f'AT{changeRow}'].value = imp

            if FINISH_PHASE == phase:
                if not impSheet['D6'].value:
                    print("need to implement pop up for importing an empty finish phase")
                else:
                    dashboard.active[f'E{changeRow}'].value = impSheet['D6'].value
                    dashboard.active[f'F{changeRow}'].value = impSheet['N26'].value
                    dashboard.active[f'G{changeRow}'].value = impSheet['O26'].value
                    dashboard.active[f'H{changeRow}'].value = impSheet['P26'].value
                    dashboard.active[f'I{changeRow}'].value = impSheet['D34'].value
                    dashboard.active[f'J{changeRow}'].value = impSheet['E34'].value
                    dashboard.active[f'K{changeRow}'].value = impSheet['F34'].value
                    dashboard.active[f'L{changeRow}'].value = impSheet['D35'].value
                    dashboard.active[f'M{changeRow}'].value = impSheet['E35'].value
                    dashboard.active[f'N{changeRow}'].value = impSheet['F35'].value
                    dashboard.active[f'O{changeRow}'].value = impSheet['D36'].value
                    dashboard.active[f'P{changeRow}'].value = impSheet['E36'].value
                    dashboard.active[f'Q{changeRow}'].value = impSheet['F36'].value
                    dashboard.active[f'U{changeRow}'].value = impSheet['N21'].value
                    dashboard.active[f'V{changeRow}'].value = impSheet['O21'].value
                    dashboard.active[f'W{changeRow}'].value = impSheet['P21'].value
                    dashboard.active[f'AA{changeRow}'].value = impSheet['D59'].value
                    dashboard.active[f'AB{changeRow}'].value = impSheet['E59'].value
                    dashboard.active[f'AC{changeRow}'].value = impSheet['F59'].value
                    dashboard.active[f'AD{changeRow}'].value = impSheet['D34'].value
                    dashboard.active[f'AE{changeRow}'].value = impSheet['L34'].value
                    dashboard.active[f'AF{changeRow}'].value = impSheet['M34'].value
                    dashboard.active[f'AG{changeRow}'].value = \
                        impSheet['D42'].value + impSheet['D53'].value
                    dashboard.active[f'AH{changeRow}'].value = \
                        impSheet['E42'].value + impSheet['E53'].value
                    dashboard.active[f'AI{changeRow}'].value = \
                        impSheet['F42'].value + impSheet['F53'].value
                    dashboard.active[f'AJ{changeRow}'].value = \
                        impSheet['D44'].value + impSheet['D46'].value + \
                        impSheet['D55'].value + impSheet['D57'].value
                    dashboard.active[f'AK{changeRow}'].value = \
                        impSheet['E44'].value + impSheet['E46'].value + \
                        impSheet['E55'].value + impSheet['E57'].value
                    dashboard.active[f'AL{changeRow}'].value = \
                        impSheet['F44'].value + impSheet['F46'].value + \
                        impSheet['F55'].value + impSheet['F57'].value
                    dashboard.active[f'AM{changeRow}'].value = \
                        impSheet['D45'].value + impSheet['D56'].value
                    dashboard.active[f'AN{changeRow}'].value = \
                        impSheet['E45'].value + impSheet['E56'].value
                    dashboard.active[f'AO{changeRow}'].value = \
                        impSheet['F45'].value + impSheet['F56'].value
                    dashboard.active[f'AP{changeRow}'].value = impSheet['D73'].value
                    dashboard.active[f'AQ{changeRow}'].value = impSheet['E73'].value
                    dashboard.active[f'AR{changeRow}'].value = impSheet['F73'].value

    dashboard.save(dashDir)


def append_default_dashboard(impDir, phase):
    dashDir = r"C:\Users\Julian.Kizanis\SAV Digital Environments\SAV - Documents\Departments\Accounting\Job Costing\00 Master Job Costing Sheet\Job Costing_Master_Dashboard.xlsx"
    # dashDir = "testing.xlsx"
    dashboard = load_workbook(filename=dashDir, read_only=False, data_only=True)
    for imp in impDir:
        impBook = load_workbook(filename=imp, read_only=False, data_only=True)
        impSheet = impBook['Data-Calculations']

        global changeRow
        changeRow = 0
        for cell in dashboard.active['Q']:
            print(cell.value)
            if cell.value == imp and (phase == ROUGH_PHASE or dashboard.active[f'E{cell.row}']):
                global openDataSheet
                openDataSheet = f"Name:{impSheet['D2'].value}\nLocation:{impSheet['D3'].value}\nPM:{impSheet['D4'].value}\nDirectory:{imp}"
                changeRow = FirstFrame.duplicate_import(cell.row, imp)
            if cell.value:
                lastRow = cell.row

        if changeRow == 0:
            changeRow = lastRow + 1
            print(f"ChangeRow:{changeRow}")
        if changeRow > -1:
            if not impSheet['D5'].value:
                print("need to implement pop up for importing an empty finish phase")
            else:
                dashboard.active[f'A{changeRow}'].value = impSheet['D2'].value
                dashboard.active[f'B{changeRow}'].value = impSheet['D4'].value
                dashboard.active[f'C{changeRow}'].value = impSheet['D5'].value
                dashboard.active[f'D{changeRow}'].value = impSheet['D47'].value
                dashboard.active[f'E{changeRow}'].value = impSheet['E47'].value
                dashboard.active[f'F{changeRow}'].value = impSheet['F73'].value
                dashboard.active[f'G{changeRow}'].value = impSheet['K47'].value
                dashboard.active[f'Q{changeRow}'].value = imp

            if FINISH_PHASE == phase:
                if not impSheet['D5'].value:
                    print("need to implement pop up for importing an empty finish phase")
                else:
                    dashboard.active[f'H{changeRow}'].value = impSheet['D6'].value
                    dashboard.active[f'I{changeRow}'].value = impSheet['H59'].value
                    dashboard.active[f'J{changeRow}'].value = impSheet['P26'].value
                    dashboard.active[f'K{changeRow}'].value = impSheet['F34'].value
                    dashboard.active[f'L{changeRow}'].value = impSheet['F35'].value
                    dashboard.active[f'M{changeRow}'].value = impSheet['D36'].value
                    dashboard.active[f'N{changeRow}'].value = impSheet['E36'].value
                    dashboard.active[f'O{changeRow}'].value = impSheet['F36'].value

    dashboard.save(dashDir)


# Define File Drop Target class
class FileDropTarget(wx.FileDropTarget):
    """ This object implements Drop Target functionality for Files """

    def __init__(self, obj):
        """ Initialize the Drop Target, passing in the Object Reference to
        indicate what should receive the dropped files """
        # Initialize the wxFileDropTarget Object
        wx.FileDropTarget.__init__(self)
        # Store the Object Reference for dropped files
        self.obj = obj

    def OnDropFiles(self, x, y, filenames):
        """ Implement File Drop """
        # For Demo purposes, this function appends a list of the files dropped at the end of the widget's text
        # Move Insertion Point to the end of the widget's text
        self.obj.SetInsertionPointEnd()
        # append a list of the file names dropped
        global ImportFiles
        dupCheck = False
        for file in filenames:
            for iFile in ImportFiles:
                if file == iFile:
                    dupCheck = True
            if dupCheck == False:
                self.obj.WriteText(file + '\n')
                ImportFiles.append(file)
            else:
                print("Removed duplicate import file!")
                dupCheck = False
        self.obj.WriteText('\n')


class FirstFrame(wx.Frame):
    def __init__(self, *args, **kwds):
        # begin wxGlade: FirstFrame.__init__
        kwds["style"] = kwds.get("style", 0) | wx.DEFAULT_FRAME_STYLE
        wx.Frame.__init__(self, *args, **kwds)
        self.SetSize((600, 428))
        self.button_4 = wx.FilePickerCtrl(self)
        # self.button_4.Bind(wx.EVT_FILEPICKER_CHANGED, self.on_choose_file)
        self.text_ctrl_1 = wx.TextCtrl(self, wx.ID_ANY, "", style=wx.TE_MULTILINE | wx.TE_READONLY)
        dropTarget = FileDropTarget(self.text_ctrl_1)
        # Link the Drop Target Object to the Text Control
        self.text_ctrl_1.SetDropTarget(dropTarget)

        self.choice_1 = wx.Choice(self, wx.ID_ANY, choices=["Choose Phase", "Rough In", "Finish"])
        self.checkbox_3 = wx.CheckBox(self, wx.ID_ANY, "General Master Dashboard")
        self.checkbox_4 = wx.CheckBox(self, wx.ID_ANY, "Kaceys's Master Dashboard")
        self.checkbox_5 = wx.CheckBox(self, wx.ID_ANY, "Jake's Master Dashboard")
        self.panel_1 = wx.Panel(self, wx.ID_ANY)
        self.button_2 = wx.Button(self, wx.ID_ANY, "Continue")
        self.button_3 = wx.Button(self, wx.ID_ANY, "Cancel")

        self.__set_properties()
        self.__do_layout()

        self.Bind(wx.EVT_FILEPICKER_CHANGED, self.on_choose_file, self.button_4)
        self.Bind(wx.EVT_CHOICE, self.on_phase_selection, self.choice_1)
        self.Bind(wx.EVT_CHECKBOX, self.on_general_master_dashboard_checkbox, self.checkbox_3)
        self.Bind(wx.EVT_CHECKBOX, self.on_kaceys_master_dashboard_checkbox, self.checkbox_4)
        self.Bind(wx.EVT_CHECKBOX, self.on_jakes_master_dashboard_checkbox, self.checkbox_5)
        self.Bind(wx.EVT_BUTTON, self.on_continue_from_main_window, self.button_2)
        self.Bind(wx.EVT_BUTTON, self.on_cancel_program, self.button_3)

    # end wxGlade

    def __set_properties(self):
        # begin wxGlade: FirstFrame.__set_properties
        self.SetTitle("Import Project Datasheet")
        self.SetBackgroundColour(wx.Colour(255, 255, 255))
        self.choice_1.SetMinSize((102, 23))
        self.choice_1.SetSelection(0)
        self.checkbox_3.SetValue(1)
        self.checkbox_4.SetValue(1)
        self.checkbox_5.SetValue(0)

    # end wxGlade

    def __do_layout(self):
        # begin wxGlade: FirstFrame.__do_layout
        sizer_5 = wx.BoxSizer(wx.VERTICAL)
        sizer_9 = wx.GridBagSizer(0, 0)
        sizer_6 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_8 = wx.BoxSizer(wx.VERTICAL)
        sizer_11 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_15 = wx.BoxSizer(wx.VERTICAL)
        sizer_12 = wx.BoxSizer(wx.VERTICAL)
        sizer_13 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_7 = wx.BoxSizer(wx.VERTICAL)
        sizer_14 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_16 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_10 = wx.BoxSizer(wx.VERTICAL)
        label_1 = wx.StaticText(self, wx.ID_ANY, "Put instructions here")
        sizer_10.Add(label_1, 0, wx.ALL, 5)
        static_line_1 = wx.StaticLine(self, wx.ID_ANY)
        sizer_10.Add(static_line_1, 0, wx.EXPAND, 0)
        sizer_5.Add(sizer_10, 0, wx.EXPAND, 0)
        sizer_16.Add(self.button_4, 0, wx.ALL, 12)
        label_6 = wx.StaticText(self, wx.ID_ANY, "Or drag and drop files below")
        sizer_16.Add(label_6, 0, wx.ALIGN_CENTER, 0)
        sizer_7.Add(sizer_16, 0, wx.EXPAND, 0)
        sizer_14.Add(self.text_ctrl_1, 1, wx.EXPAND, 0)
        sizer_7.Add(sizer_14, 1, wx.EXPAND, 0)
        sizer_6.Add(sizer_7, 2, wx.EXPAND, 0)
        bitmap_2 = wx.StaticBitmap(self, wx.ID_ANY, wx.Bitmap(
            "C:\\Users\\Julian.Kizanis\\OneDrive - SAV Digital Environments\\Python Code\\Job Costing Program\\SAV-Company-Logo.png",
            wx.BITMAP_TYPE_ANY))
        sizer_12.Add(bitmap_2, 0, wx.BOTTOM | wx.RIGHT | wx.TOP, 12)
        sizer_13.Add(self.choice_1, 0, wx.BOTTOM | wx.LEFT, 6)
        sizer_12.Add(sizer_13, 1, wx.EXPAND, 0)
        sizer_8.Add(sizer_12, 0, wx.EXPAND, 0)
        sizer_15.Add(self.checkbox_3, 0, wx.LEFT | wx.RIGHT | wx.TOP, 6)
        sizer_15.Add(self.checkbox_4, 0, wx.LEFT | wx.RIGHT | wx.TOP, 6)
        sizer_15.Add(self.checkbox_5, 0, wx.LEFT | wx.RIGHT | wx.TOP, 6)
        sizer_11.Add(sizer_15, 1, wx.EXPAND, 0)
        sizer_8.Add(sizer_11, 1, wx.EXPAND, 0)
        sizer_6.Add(sizer_8, 0, wx.EXPAND | wx.LEFT, 6)
        sizer_5.Add(sizer_6, 1, wx.EXPAND, 0)
        sizer_9.Add(self.panel_1, (0, 0), (1, 1), wx.EXPAND, 0)
        sizer_9.Add(self.button_2, (0, 1), (1, 1), wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL, 6)
        sizer_9.Add(self.button_3, (0, 2), (1, 1), wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL, 6)
        sizer_5.Add(sizer_9, 0, wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL | wx.EXPAND, 12)
        self.SetSizer(sizer_5)
        self.Layout()

    # end wxGlade

    def on_choose_file(self, event):  # wxGlade: FirstFrame.<event_handler>
        global ImportFiles
        dupCheck = False
        file = self.button_4.GetPath()
        for iFile in ImportFiles:
            if file == iFile:
                dupCheck = True

        if dupCheck == False:
            ImportFiles.append(file)
            self.text_ctrl_1.WriteText(file + '\n')
        else:
            print("Removed duplicate import file!")
            dupCheck = False
        event.Skip()

    def on_phase_selection(self, event):  # wxGlade: FirstFrame.<event_handler>
        print(self.choice_1.GetSelection())
        event.Skip()

    def on_general_master_dashboard_checkbox(self, event):  # wxGlade: FirstFrame.<event_handler>
        print(self.checkbox_3.GetValue())
        event.Skip()

    def on_kaceys_master_dashboard_checkbox(self, event):  # wxGlade: FirstFrame.<event_handler>
        print(self.checkbox_4.GetValue())
        event.Skip()

    def on_jakes_master_dashboard_checkbox(self, event):  # wxGlade: FirstFrame.<event_handler>
        print(self.checkbox_5.GetValue())
        event.Skip()

    def on_continue_from_main_window(self, event):  # wxGlade: FirstFrame.<event_handler>
        if self.choice_1.GetSelection() == 0:
            wx.MessageBox("Please choose a phase.", "Error", wx.OK | wx.ICON_INFORMATION)
        elif not ImportFiles:
            wx.MessageBox("Please choose a file to import.", "Error", wx.OK | wx.ICON_INFORMATION)
        else:
            if self.checkbox_3.GetValue():
                append_default_dashboard(ImportFiles, self.choice_1.GetSelection())
            if self.checkbox_4.GetValue():
                append_kaceys_dashboard(ImportFiles, self.choice_1.GetSelection())
        event.Skip()

    def on_cancel_program(self, event):  # wxGlade: FirstFrame.<event_handler>
        self.Destroy()
        event.Skip()

    def duplicate_import(self, row, imp):
        frame = Datasheet_already_imported_frame(self)
        frame.show()


# end of class FirstFrame

class Datasheet_already_imported_dialog(wx.Dialog):
    def __init__(self, *args, **kwds):
        # begin wxGlade: Datasheet_already_imported_dialog.__init__
        kwds["style"] = kwds.get("style", 0) | wx.DEFAULT_DIALOG_STYLE
        wx.Dialog.__init__(self, *args, **kwds)
        self.text_ctrl_already_imported = wx.TextCtrl(self, wx.ID_ANY, \
                                                      "{openDataSheet} has already been imported by {importedBy} on {importedDate}. If you would  like to import the project as a new project, select \"Duplicate\". If you want to refresh the exsisting data, select \"Replace\".", \
                                                      style=wx.BORDER_NONE | wx.TE_MULTILINE | wx.TE_READONLY)
        self.panel_2 = wx.Panel(self, wx.ID_ANY)
        self.button_6 = wx.Button(self, wx.ID_ANY, "Duplicate")
        self.button_1 = wx.Button(self, wx.ID_ANY, "Replace")
        self.button_5 = wx.Button(self, wx.ID_ANY, "Back")

        self.__set_properties()
        self.__do_layout()

        self.Bind(wx.EVT_TEXT, self.text_ctrl_openDataSheet, self.text_ctrl_already_imported)
        self.Bind(wx.EVT_BUTTON, self.on_duplicate, self.button_6)
        self.Bind(wx.EVT_BUTTON, self.on_replace, self.button_1)
        self.Bind(wx.EVT_BUTTON, self.on_back, self.button_5)

    # end wxGlade

    def __set_properties(self):
        # begin wxGlade: Datasheet_already_imported_dialog.__set_properties
        self.SetTitle("dialog_2")
        self.text_ctrl_already_imported.SetBackgroundColour(wx.Colour(255, 255, 255))

    # end wxGlade

    def __do_layout(self):
        # begin wxGlade: Datasheet_already_imported_dialog.__do_layout
        sizer_1 = wx.BoxSizer(wx.VERTICAL)
        sizer_2 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_1.Add(self.text_ctrl_already_imported, 0, wx.ALL | wx.EXPAND, 12)
        sizer_2.Add(self.panel_2, 1, 0, 0)
        sizer_2.Add(self.button_6, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_2.Add(self.button_1, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_2.Add(self.button_5, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_1.Add(sizer_2, 1, wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL | wx.EXPAND | wx.FIXED_MINSIZE, 1)
        self.SetSizer(sizer_1)
        sizer_1.Fit(self)
        self.Layout()

    # end wxGlade

    def text_ctrl_openDataSheet(self, event):  # wxGlade: Datasheet_already_imported_dialog.<event_handler>
        print("Event handler 'text_ctrl_openDataSheet' not implemented!")
        event.Skip()

    def on_duplicate(self, event):  # wxGlade: Datasheet_already_imported_dialog.<event_handler>
        print("Event handler 'on_duplicate' not implemented!")
        event.Skip()

    def on_replace(self, event):  # wxGlade: Datasheet_already_imported_dialog.<event_handler>
        print("Event handler 'on_replace' not implemented!")
        event.Skip()

    def on_back(self, event):  # wxGlade: Datasheet_already_imported_dialog.<event_handler>
        print("Event handler 'on_back' not implemented!")
        event.Skip()


# end of class Datasheet_open_frame

class Datasheet_already_imported_frame(wx.Frame):
    def __init__(self, *args, **kwds):
        # begin wxGlade: Datasheet_already_imported_frame.__init__
        kwds["style"] = kwds.get("style", 0) | wx.DEFAULT_FRAME_STYLE | wx.STAY_ON_TOP
        wx.Frame.__init__(self, *args, **kwds)
        self.SetSize((350, 275))
        self.text_ctrl_open_datasheet = wx.TextCtrl(self, wx.ID_ANY,
                                                    "{openDataSheet}\nHas already been imported. If you would  like to import the project as a new project, select \"Duplicate\". If you want to refresh the exsisting data, select \"Replace\".",
                                                    style=wx.BORDER_NONE | wx.TE_MULTILINE | wx.TE_READONLY)
        self.panel_2 = wx.Panel(self, wx.ID_ANY)
        self.button_6 = wx.Button(self, wx.ID_ANY, "Duplicate")
        self.button_1 = wx.Button(self, wx.ID_ANY, "Replace")
        self.button_5 = wx.Button(self, wx.ID_ANY, "Back")

        self.__set_properties()
        self.__do_layout()

        self.Bind(wx.EVT_TEXT, self.text_ctrl_openDataSheet, self.text_ctrl_open_datasheet)
        self.Bind(wx.EVT_BUTTON, self.on_duplicate, self.button_6)
        self.Bind(wx.EVT_BUTTON, self.on_replace, self.button_1)
        self.Bind(wx.EVT_BUTTON, self.on_back, self.button_5)

    # end wxGlade

    def __set_properties(self):
        # begin wxGlade: Datasheet_already_imported_frame.__set_properties
        self.SetTitle("frame_2")
        self.SetBackgroundColour(wx.Colour(255, 255, 255))
        self.text_ctrl_open_datasheet.SetBackgroundColour(wx.Colour(255, 255, 255))

    # end wxGlade

    def __do_layout(self):
        # begin wxGlade: Datasheet_already_imported_frame.__do_layout
        sizer_1 = wx.BoxSizer(wx.VERTICAL)
        sizer_2 = wx.BoxSizer(wx.HORIZONTAL)
        sizer_1.Add(self.text_ctrl_open_datasheet, 0, wx.ALL | wx.EXPAND, 12)
        sizer_2.Add(self.panel_2, 1, 0, 0)
        sizer_2.Add(self.button_6, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_2.Add(self.button_1, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_2.Add(self.button_5, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_1.Add(sizer_2, 1, wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL | wx.EXPAND | wx.FIXED_MINSIZE, 1)
        self.SetSizer(sizer_1)
        self.Layout()

    # end wxGlade

    def text_ctrl_openDataSheet(self, event):  # wxGlade: Datasheet_already_imported_frame.<event_handler>
        print("Event handler 'text_ctrl_openDataSheet' not implemented!")
        event.Skip()

    def on_duplicate(self, event):  # wxGlade: Datasheet_already_imported_frame.<event_handler>
        global changeRow
        changeRow = 0
        self.Destroy()
        event.Skip()

    def on_replace(self, event):  # wxGlade: Datasheet_already_imported_frame.<event_handler>
        self.destroy()
        event.Skip()

    def on_back(self, event):  # wxGlade: Datasheet_already_imported_frame.<event_handler>
        global changeRow
        changeRow = -1
        self.Destroy()
        event.Skip()


# end of class Datasheet_already_imported_frame

class are_you_sure_replace_dialog(wx.Dialog):
    def __init__(self, *args, **kwds):
        # begin wxGlade: are_you_sure_replace_dialog.__init__
        kwds["style"] = kwds.get("style", 0) | wx.DEFAULT_DIALOG_STYLE
        wx.Dialog.__init__(self, *args, **kwds)
        self.panel_2 = wx.Panel(self, wx.ID_ANY)
        self.button_1 = wx.Button(self, wx.ID_ANY, "Replace")
        self.button_5 = wx.Button(self, wx.ID_ANY, "Back")

        self.__set_properties()
        self.__do_layout()

        self.Bind(wx.EVT_BUTTON, self.on_replace, self.button_1)
        self.Bind(wx.EVT_BUTTON, self.on_back, self.button_5)

    # end wxGlade

    def __set_properties(self):
        # begin wxGlade: are_you_sure_replace_dialog.__set_properties
        self.SetTitle("dialog")
        _icon = wx.NullIcon
        _icon.CopyFromBitmap(wx.Bitmap(
            "C:\\Users\\Julian.Kizanis\\OneDrive - SAV Digital Environments\\Python Code\\Job Costing Program\\SAV-Social-Profile.jpg",
            wx.BITMAP_TYPE_ANY))
        self.SetIcon(_icon)

    # end wxGlade

    def __do_layout(self):
        # begin wxGlade: are_you_sure_replace_dialog.__do_layout
        sizer_1 = wx.BoxSizer(wx.VERTICAL)
        sizer_2 = wx.BoxSizer(wx.HORIZONTAL)
        label_2 = wx.StaticText(self, wx.ID_ANY,
                                "Are you Sure you want to replace/overwrite the project? The old data will not be saved.")
        label_2.Wrap(300)
        sizer_1.Add(label_2, 0, wx.ALL, 12)
        sizer_2.Add(self.panel_2, 1, 0, 0)
        sizer_2.Add(self.button_1, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_2.Add(self.button_5, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_1.Add(sizer_2, 1, wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL | wx.EXPAND | wx.FIXED_MINSIZE, 1)
        self.SetSizer(sizer_1)
        sizer_1.Fit(self)
        self.Layout()

    # end wxGlade

    def on_replace(self, event):  # wxGlade: are_you_sure_replace_dialog.<event_handler>
        print("Event handler 'on_replace' not implemented!")
        event.Skip()

    def on_back(self, event):  # wxGlade: are_you_sure_replace_dialog.<event_handler>
        print("Event handler 'on_back' not implemented!")
        event.Skip()


# end of class are_you_sure_replace_dialog

class are_you_sure_duplicate_dialog(wx.Dialog):
    def __init__(self, *args, **kwds):
        # begin wxGlade: are_you_sure_duplicate_dialog.__init__
        kwds["style"] = kwds.get("style", 0) | wx.DEFAULT_DIALOG_STYLE
        wx.Dialog.__init__(self, *args, **kwds)
        self.panel_2 = wx.Panel(self, wx.ID_ANY)
        self.button_1 = wx.Button(self, wx.ID_ANY, "Duplicate")
        self.button_5 = wx.Button(self, wx.ID_ANY, "Back")

        self.__set_properties()
        self.__do_layout()

        self.Bind(wx.EVT_BUTTON, self.on_duplicate, self.button_1)
        self.Bind(wx.EVT_BUTTON, self.on_back, self.button_5)

    # end wxGlade

    def __set_properties(self):
        # begin wxGlade: are_you_sure_duplicate_dialog.__set_properties
        self.SetTitle("dialog_1")

    # end wxGlade

    def __do_layout(self):
        # begin wxGlade: are_you_sure_duplicate_dialog.__do_layout
        sizer_1 = wx.BoxSizer(wx.VERTICAL)
        sizer_2 = wx.BoxSizer(wx.HORIZONTAL)
        label_2 = wx.StaticText(self, wx.ID_ANY, "Are you Sure you want to add the project as a duplicate?")
        label_2.Wrap(300)
        sizer_1.Add(label_2, 0, wx.ALL, 12)
        sizer_2.Add(self.panel_2, 1, 0, 0)
        sizer_2.Add(self.button_1, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_2.Add(self.button_5, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_1.Add(sizer_2, 1, wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL | wx.EXPAND | wx.FIXED_MINSIZE, 1)
        self.SetSizer(sizer_1)
        sizer_1.Fit(self)
        self.Layout()

    # end wxGlade

    def on_duplicate(self, event):  # wxGlade: are_you_sure_duplicate_dialog.<event_handler>
        print("Event handler 'on_duplicate' not implemented!")
        event.Skip()

    def on_back(self, event):  # wxGlade: are_you_sure_duplicate_dialog.<event_handler>
        print("Event handler 'on_back' not implemented!")
        event.Skip()


# end of class are_you_sure_duplicate_dialog

class success_frame(wx.Frame):
    def __init__(self, *args, **kwds):
        # begin wxGlade: success_frame.__init__
        kwds["style"] = kwds.get("style", 0) | wx.DEFAULT_FRAME_STYLE | wx.STAY_ON_TOP
        wx.Frame.__init__(self, *args, **kwds)
        self.SetSize((350, 150))
        self.panel_2 = wx.Panel(self, wx.ID_ANY)
        self.button_5 = wx.Button(self, wx.ID_ANY, "Okay")

        self.__set_properties()
        self.__do_layout()

        self.Bind(wx.EVT_BUTTON, self.on_okay, self.button_5)

    # end wxGlade

    def __set_properties(self):
        # begin wxGlade: success_frame.__set_properties
        self.SetTitle("frame_2")
        self.SetBackgroundColour(wx.Colour(255, 255, 255))

    # end wxGlade

    def __do_layout(self):
        # begin wxGlade: success_frame.__do_layout
        sizer_1 = wx.BoxSizer(wx.VERTICAL)
        sizer_2 = wx.BoxSizer(wx.HORIZONTAL)
        label_2 = wx.StaticText(self, wx.ID_ANY, "The project was successfully imported!")
        label_2.Wrap(300)
        sizer_1.Add(label_2, 0, wx.ALL, 12)
        sizer_2.Add(self.panel_2, 1, 0, 0)
        sizer_2.Add(self.button_5, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_1.Add(sizer_2, 1, wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL | wx.EXPAND | wx.FIXED_MINSIZE, 1)
        self.SetSizer(sizer_1)
        self.Layout()

    # end wxGlade

    def on_okay(self, event):  # wxGlade: success_frame.<event_handler>
        print("Event handler 'on_okay' not implemented!")
        event.Skip()

    # end of class success_frame

    # class error_frame(wx.Frame):
    #     def __init__(self, *args, **kwds):
    #         # begin wxGlade: error_frame.__init__
    #         kwds["style"] = kwds.get("style", 0) | wx.DEFAULT_FRAME_STYLE | wx.STAY_ON_TOP
    #         wx.Frame.__init__(self, *args, **kwds)
    #         self.SetSize((350, 150))
    #         self.panel_2 = wx.Panel(self, wx.ID_ANY)
    #         self.button_5 = wx.Button(self, wx.ID_ANY, "Okay")
    #
    #         self.__set_properties()
    #         self.__do_layout()
    #
    #         self.Bind(wx.EVT_BUTTON, self.on_okay, self.button_5)

    # end wxGlade

    def __set_properties(self):
        # begin wxGlade: error_frame.__set_properties
        self.SetTitle("frame_2")
        self.SetBackgroundColour(wx.Colour(255, 255, 255))

    # end wxGlade

    def __do_layout(self):
        # begin wxGlade: error_frame.__do_layout
        sizer_1 = wx.BoxSizer(wx.VERTICAL)
        sizer_2 = wx.BoxSizer(wx.HORIZONTAL)
        label_2 = wx.StaticText(self, wx.ID_ANY, "An unexpected error has occurred!")
        label_2.Wrap(300)
        sizer_1.Add(label_2, 0, wx.ALL, 12)
        sizer_2.Add(self.panel_2, 1, 0, 0)
        sizer_2.Add(self.button_5, 0, wx.ALIGN_BOTTOM | wx.ALL | wx.FIXED_MINSIZE, 12)
        sizer_1.Add(sizer_2, 1, wx.ALIGN_BOTTOM | wx.ALIGN_RIGHT | wx.ALL | wx.EXPAND | wx.FIXED_MINSIZE, 1)
        self.SetSizer(sizer_1)
        self.Layout()

    # end wxGlade

    def on_okay(self, event):  # wxGlade: error_frame.<event_handler>
        print("Event handler 'on_okay' not implemented!")
        event.Skip()


# end of class error_frame

class MyApp(wx.App):
    def OnInit(self):
        self.frame = FirstFrame(None, wx.ID_ANY, "")
        self.SetTopWindow(self.frame)
        self.frame.Show()
        return True


# end of class MyApp

if __name__ == "__main__":
    ImportProjectDatasheets = MyApp(0)
    ImportProjectDatasheets.MainLoop()
